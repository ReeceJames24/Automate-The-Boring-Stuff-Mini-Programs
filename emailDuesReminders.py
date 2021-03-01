#emailDuesReminders.py - send emails based on payment status in duesRecords.xlsx

import openpyxl, smtplib, sys

#TO DO: Open Spreadsheet File
wb = openpyxl.load_workbook('/Users/chummi/Desktop/duesRecords.xlsx')
sheet = wb['Sheet1']
lastCol = sheet.max_column
lastRow = sheet.max_row
latestMonth = sheet.cell(row=1, column=lastCol).value


#TO DO: Check payment status
unpaidMembers = {}
for i in range (2, lastRow + 1): # +1 to include last row in the loop
    paymentStatus = sheet.cell(row = i, column = lastCol).value
    if paymentStatus != 'paid':
        name = sheet.cell(row = i, column = 1).value
        email = sheet.cell(row = i, column = 2).value
        unpaidMembers[name] = email # setting key value pairs for the unpaid members
        
for i in range(1, 10001):
    if i == 10000:
        print('Initializing.....')

#TO DO: Log-in email
conn = smtplib.SMTP('smtp.gmail.com', 587)
conn.ehlo()
conn.starttls()
passw = input('Input password:')
conn.login('ndcrisologo@gmail.com', passw)


#TO DO: Send emails to unpaidMembers{}
for name, email in unpaidMembers.items():
    body = 'Subject: %s dues unpaid. \n\n Dear %s, \nRecords show that you have not paid dues for the month of %s. Click link below to start: \nm.me/boscoffeedaily. \nPay up or I will punch you. Thank you! \n\nRegards, \nChummi.' % (latestMonth, name, latestMonth)
    print('Sending email to %s....' % email)
    sendmailStatus = conn.sendmail('ndcrisologo@gmail.com', email, body)

    if sendmailStatus != {}: #since sendmail returns blank dictionary if no error occurs
        print('There was an error in sending the email to %s' % email)

conn.quit()
    


    


